#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
API Document Retrieval V4 - 能率需求清单智能匹配（支持国家平台/BPM集成）
优化匹配逻辑：
1. 开发分类 → 流向自动推断
   - 702/703/704/705: 服务端接口 → XX→T100 (OA/BPM等外部系统推送数据到T100)
   - 710/711: 客户端接口 → T100→XX (T100推送数据到OA/BPM等外部系统)
   - 202: 基本资料作业 → 根据需求描述推断
2. 程序代号 + 推断流向 → 匹配接口清单"流程名称"
3. 支持 .xls 和 .xlsx 格式
"""

import sys
import os
from pathlib import Path
from datetime import datetime
import re
import shutil

# 尝试导入 openpyxl (xlsx)
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 尝试导入 xlrd (xls)
try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

# 尝试导入 xlwt (写入xls)
try:
    import xlwt
    HAS_XLWT = True
except ImportError:
    HAS_XLWT = False

# 配置
WORKSPACE_DIR = Path(__file__).parent.parent.parent.parent.parent
API_MAPPING_FILE = WORKSPACE_DIR / "需求对照工具表" / "【2026】T100&OA类系统集成清单字段对照表.xlsx"
ONLINE_DOC_URL = "https://docs.qq.com/sheet/DSmdhek1TT3BPenRa?tab=BB08J2"


class NenglvMatcherV4:
    """能率需求清单智能匹配器 V4"""
    
    # 开发分类 → 流向映射规则
    FLOW_DIRECTION_RULES = {
        # 服务端接口：外部系统 → T100
        '702': 'XX→T100',  # 服务端_生单接口
        '703': 'XX→T100',  # 服务端_更新接口
        '704': 'XX→T100',  # 服务端_更新接口(审核)
        '705': 'XX→T100',  # 服务端_其他接口
        # 客户端接口：T100 → 外部系统
        '710': 'T100→XX',  # 客户端_推送接口
        '711': 'T100→XX',  # 客户端_其他推送
    }
    
    def __init__(self):
        self.api_data = []  # 接口清单数据
        self.load_api_mapping()
    
    def load_api_mapping(self):
        """加载API对照清单 - 接口清单页签"""
        if not API_MAPPING_FILE.exists():
            print(f"警告: 找不到本地API对照清单文件: {API_MAPPING_FILE}")
            print(f"\n建议使用在线文档: {ONLINE_DOC_URL}")
            sys.exit(1)
        
        print(f"正在加载API对照清单: {API_MAPPING_FILE.name}")
        
        try:
            wb = openpyxl.load_workbook(API_MAPPING_FILE, data_only=True, read_only=True)
        except Exception as e:
            print(f"警告: 使用 read_only 模式读取失败: {e}")
            try:
                wb = openpyxl.load_workbook(API_MAPPING_FILE, data_only=False, read_only=True)
            except Exception as e2:
                print(f"错误: 无法读取文件: {e2}")
                sys.exit(1)
        
        # 查找"接口清单"页签
        interface_sheet = None
        for sheet_name in wb.sheetnames:
            if '接口清单' in sheet_name:
                interface_sheet = wb[sheet_name]
                print(f"找到接口清单页签: {sheet_name}")
                break
        
        if not interface_sheet:
            print(f"未找到'接口清单'页签，使用第一个页签")
            interface_sheet = wb.active
        
        # 读取表头（第一行）
        headers = []
        first_row = list(interface_sheet.iter_rows(min_row=1, max_row=1, values_only=False))[0]
        for col_num, cell in enumerate(first_row, 1):
            if cell.value:
                headers.append({
                    'col': col_num,
                    'name': str(cell.value).strip()
                })
        
        print(f"表头字段: {[h['name'] for h in headers[:8]]}")
        
        # 读取所有行数据
        for row_num, row in enumerate(interface_sheet.iter_rows(min_row=2, values_only=False), 2):
            row_data = {'_row_num': row_num, '_sheet': interface_sheet.title}
            for header in headers:
                cell = row[header['col'] - 1]
                row_data[header['name']] = str(cell.value).strip() if cell.value else ''
            
            # 只保留有流程名称的行
            if row_data.get('流程名称', ''):
                self.api_data.append(row_data)
        
        print(f"已加载 {len(self.api_data)} 条接口记录\n")
        wb.close()
    
    def infer_flow_direction(self, dev_class):
        """根据开发分类推断流向"""
        if not dev_class:
            return ''
        
        # 提取开发分类代码（数字部分）
        match = re.search(r'(\d{3})', str(dev_class))
        if match:
            class_code = match.group(1)
            return self.FLOW_DIRECTION_RULES.get(class_code, '')
        
        return ''
    
    def normalize_flow_direction(self, flow_dir):
        """标准化流向格式"""
        if not flow_dir:
            return ''
        
        flow = str(flow_dir).upper().replace(' ', '').replace('--', '→').replace('->', '→')
        
        # 统一格式
        if 'ERP' in flow and 'OA' in flow:
            if flow.startswith('ERP') or flow.startswith('T100'):
                return 'T100→OA'
            else:
                return 'OA→T100'
        elif 'XX→T100' in flow or '→T100' in flow or 'OA→ERP' in flow:
            return 'XX→T100'
        elif 'T100→XX' in flow or 'T100→' in flow or 'ERP→OA' in flow:
            return 'T100→XX'
        
        return flow
    
    def extract_prog_code(self, text):
        """从文本中提取程序代号（如 apmt200）"""
        if not text:
            return ''
        match = re.search(r'([a-z]{2,4}\d{3})', str(text).lower())
        return match.group(1) if match else ''
    
    def match_interface(self, requirement):
        """根据需求匹配接口清单"""
        matches = []
        
        # 提取关键信息
        dev_class = str(requirement.get('开发分类', '') or requirement.get('开发分类(*)', ''))
        prog_code = str(requirement.get('程序代号', '') or requirement.get('程序代号(*)', '')).lower().strip()
        req_name = str(requirement.get('程序名称', '') or requirement.get('程序名称(*)', ''))
        spec_desc = str(requirement.get('规格说明', '') or requirement.get('规格说明(*)', ''))
        
        if not prog_code:
            return matches
        
        # 推断流向
        inferred_flow = self.infer_flow_direction(dev_class)
        requirement['_inferred_flow'] = inferred_flow
        
        print(f"  程序代号: {prog_code}, 开发分类: {dev_class}, 推断流向: {inferred_flow}")
        
        for api_row in self.api_data:
            flow_name = str(api_row.get('流程名称', ''))
            api_direction = str(api_row.get('接口方向', ''))
            
            if not flow_name:
                continue
            
            score = 0
            reasons = []
            
            # 规则1: 程序代号精确匹配（权重 60%）
            api_prog_code = self.extract_prog_code(flow_name)
            if api_prog_code == prog_code:
                score += 60
                reasons.append(f"程序代号精确匹配: {prog_code}")
            elif prog_code in flow_name.lower():
                score += 40
                reasons.append(f"程序代号包含匹配: {prog_code}")
            
            # 规则2: 流向匹配（权重 30%）
            if inferred_flow and score > 0:
                normalized_api_dir = self.normalize_flow_direction(api_direction)
                
                if inferred_flow == 'T100→XX':
                    # 期望 T100→OA 或 ERP→OA
                    if 'ERP→OA' in api_direction or 'T100→OA' in api_direction or 'ERP-->OA' in api_direction:
                        score += 30
                        reasons.append(f"流向匹配: {inferred_flow} ↔ {api_direction}")
                    elif 'OA→ERP' in api_direction or 'OA→T100' in api_direction or 'OA-->ERP' in api_direction:
                        # 方向相反，降低匹配度
                        score -= 20
                        reasons.append(f"流向相反: {inferred_flow} vs {api_direction}")
                
                elif inferred_flow == 'XX→T100':
                    # 期望 OA→T100 或 OA→ERP
                    if 'OA→ERP' in api_direction or 'OA→T100' in api_direction or 'OA-->ERP' in api_direction:
                        score += 30
                        reasons.append(f"流向匹配: {inferred_flow} ↔ {api_direction}")
                    elif 'ERP→OA' in api_direction or 'T100→OA' in api_direction or 'ERP-->OA' in api_direction:
                        # 方向相反，降低匹配度
                        score -= 20
                        reasons.append(f"流向相反: {inferred_flow} vs {api_direction}")
            
            # 规则3: 需求名称辅助匹配（权重 10%）
            if score > 0 and req_name:
                # 提取业务关键词
                req_keywords = self.extract_business_keywords(req_name)
                flow_keywords = self.extract_business_keywords(flow_name)
                
                common_keywords = set(req_keywords) & set(flow_keywords)
                if common_keywords:
                    keyword_score = min(len(common_keywords) * 5, 10)
                    score += keyword_score
                    reasons.append(f"业务关键词匹配: {', '.join(common_keywords)}")
            
            # 规则4: 规格说明中的程序代号提及（权重 5%）
            if prog_code in spec_desc.lower() and score > 0:
                score += 5
                reasons.append("规格说明中提及程序代号")
            
            # 记录匹配结果
            if score >= 40:  # 阈值
                matches.append({
                    'api': api_row,
                    'score': min(score, 100),
                    'reason': ' | '.join(reasons),
                    'match_type': '流程名称匹配'
                })
        
        # 按匹配度排序
        matches.sort(key=lambda x: x['score'], reverse=True)
        
        return matches
    
    def extract_business_keywords(self, text):
        """提取业务关键词"""
        if not text:
            return []
        
        # 常见业务关键词
        keywords = [
            '员工', '部门', '料号', '客户', '供应商', '采购', '销售', '订单',
            '核价', '请购', '入库', '出库', '库存', '杂收', '杂发', '付款',
            '收款', '报销', '借款', '预算', '资产', '理财', '发票', '核销',
            '申请', '维护', '查询', '新增', '修改', '删除', '审核', '送签',
            '推送', '同步', '抛转', '回写', '更新', '创建', '生成'
        ]
        
        text_lower = str(text).lower()
        found = [kw for kw in keywords if kw in text_lower]
        return found
    
    def read_requirement_file(self, file_path):
        """读取需求清单文件（支持 .xls 和 .xlsx）"""
        print(f"读取需求清单: {file_path.name}")
        
        requirements = []
        headers = {}
        
        if file_path.suffix.lower() == '.xlsx':
            if not HAS_OPENPYXL:
                print("错误: 未安装 openpyxl，请运行: pip install openpyxl")
                sys.exit(1)
            
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # 读取表头
            for col_num, cell in enumerate(ws[1], 1):
                if cell.value:
                    headers[str(cell.value).strip()] = col_num
            
            # 读取数据行
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
                req_data = {'_row_num': row_num, '_file_type': 'xlsx'}
                for col_name, col_num in headers.items():
                    cell = row[col_num - 1]
                    req_data[col_name] = str(cell.value).strip() if cell.value else ''
                
                # 只保留有程序代号的行
                prog_code = req_data.get('程序代号', '') or req_data.get('程序代号(*)', '')
                if prog_code:
                    requirements.append(req_data)
        
        elif file_path.suffix.lower() == '.xls':
            if not HAS_XLRD:
                print("错误: 未安装 xlrd，请运行: pip install xlrd")
                sys.exit(1)
            
            wb = xlrd.open_workbook(file_path)
            ws = wb.sheet_by_index(0)
            
            # 读取表头
            for col_idx in range(ws.ncols):
                cell_value = ws.cell(0, col_idx).value
                if cell_value:
                    headers[str(cell_value).strip()] = col_idx
            
            # 读取数据行
            for row_idx in range(1, ws.nrows):
                req_data = {'_row_num': row_idx + 1, '_file_type': 'xls'}
                for col_name, col_idx in headers.items():
                    cell_value = ws.cell(row_idx, col_idx).value
                    req_data[col_name] = str(cell_value).strip() if cell_value else ''
                
                # 只保留有程序代号的行
                prog_code = req_data.get('程序代号', '') or req_data.get('程序代号(*)', '')
                if prog_code:
                    requirements.append(req_data)
        
        print(f"工作表: 导入模板")
        print(f"提取到 {len(requirements)} 条需求记录\n")
        return requirements, headers
    
    def generate_output_xlsx(self, requirements, results, original_file):
        """生成 xlsx 格式的匹配结果"""
        print("\n生成匹配结果文件...")
        
        # 备份原文件
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_file = original_file.parent / f"{original_file.stem}_备份_{timestamp}{original_file.suffix}"
        shutil.copy2(original_file, backup_file)
        print(f"原文件已备份: {backup_file.name}")
        
        # 创建新的工作簿
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = "需求匹配结果"
        
        # 读取原始表头
        if original_file.suffix.lower() == '.xlsx':
            wb_in = openpyxl.load_workbook(original_file, data_only=True)
            ws_in = wb_in.active
            original_headers = []
            for col_num, cell in enumerate(ws_in[1], 1):
                original_headers.append(str(cell.value) if cell.value else f'Column{col_num}')
        else:
            # 对于 .xls 文件，使用之前读取的headers
            wb_in = xlrd.open_workbook(original_file)
            ws_in = wb_in.sheet_by_index(0)
            original_headers = []
            for col_idx in range(ws_in.ncols):
                cell_value = ws_in.cell(0, col_idx).value
                original_headers.append(str(cell_value) if cell_value else f'Column{col_idx+1}')
        
        # 添加匹配结果列
        match_headers = [
            '匹配状态',
            '匹配流程名称',
            '接口方向',
            '推断流向',
            '匹配度',
            '匹配依据',
            '参考位置',
            '在线文档链接'
        ]
        
        all_headers = original_headers + match_headers
        
        # 写入表头
        for col_idx, header in enumerate(all_headers, 1):
            cell = ws_out.cell(row=1, column=col_idx, value=header)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # 写入数据
        for i, result in enumerate(results, 2):
            req = result['requirement']
            api = result['api']
            score = result['score']
            reason = result['reason']
            inferred_flow = result.get('inferred_flow', '')
            
            # 写入原始数据
            for col_idx, header in enumerate(original_headers, 1):
                value = req.get(header, '')
                ws_out.cell(row=i, column=col_idx, value=value)
            
            # 写入匹配结果
            base_col = len(original_headers) + 1
            
            if api and score >= 40:
                ws_out.cell(row=i, column=base_col, value="✅ 已匹配")
                ws_out.cell(row=i, column=base_col + 1, value=api.get('流程名称', ''))
                ws_out.cell(row=i, column=base_col + 2, value=api.get('接口方向', ''))
                ws_out.cell(row=i, column=base_col + 3, value=inferred_flow)
                ws_out.cell(row=i, column=base_col + 4, value=f"{score}%")
                ws_out.cell(row=i, column=base_col + 5, value=reason)
                ws_out.cell(row=i, column=base_col + 6, value=f"接口清单 行{api.get('_row_num', '')}")
                ws_out.cell(row=i, column=base_col + 7, value=ONLINE_DOC_URL)
            else:
                ws_out.cell(row=i, column=base_col, value="❌ 未匹配")
                ws_out.cell(row=i, column=base_col + 1, value="")
                ws_out.cell(row=i, column=base_col + 2, value="")
                ws_out.cell(row=i, column=base_col + 3, value=inferred_flow)
                ws_out.cell(row=i, column=base_col + 4, value="0%")
                ws_out.cell(row=i, column=base_col + 5, value=reason or "无匹配接口")
                ws_out.cell(row=i, column=base_col + 6, value="")
                ws_out.cell(row=i, column=base_col + 7, value=ONLINE_DOC_URL)
        
        # 设置列宽
        for col_idx in range(1, len(all_headers) + 1):
            ws_out.column_dimensions[get_column_letter(col_idx)].width = 25
        
        # 保存文件
        output_file = original_file.parent / f"{original_file.stem}_API匹配结果.xlsx"
        
        try:
            wb_out.save(output_file)
            print(f"匹配结果已保存: {output_file.name}")
            print(f"文件大小: {output_file.stat().st_size / 1024:.2f} KB\n")
        except PermissionError:
            output_file = original_file.parent / f"{original_file.stem}_API匹配结果_{timestamp}.xlsx"
            wb_out.save(output_file)
            print(f"匹配结果已保存: {output_file.name}")
            print(f"文件大小: {output_file.stat().st_size / 1024:.2f} KB\n")
        
        return output_file
    
    def process_requirement_file(self, file_path):
        """处理需求清单文件"""
        print(f"\n📖 处理需求清单: {file_path.name}")
        print("=" * 80)
        
        # 读取需求清单
        requirements, headers = self.read_requirement_file(file_path)
        
        if not requirements:
            print("未找到有效的需求记录")
            return None
        
        # 匹配接口
        print("开始匹配API...")
        print("=" * 80)
        
        results = []
        matched_count = 0
        
        for i, req in enumerate(requirements, 1):
            prog_code = req.get('程序代号(*)', '') or req.get('程序代号', '')
            req_name = req.get('程序名称(*)', '') or req.get('程序名称', '')
            dev_class = req.get('开发分类(*)', '') or req.get('开发分类', '')
            
            print(f"[{i}/{len(requirements)}] {prog_code} - {req_name[:40]}...")
            print(f"  开发分类: {dev_class}")
            
            matches = self.match_interface(req)
            
            if matches:
                best_match = matches[0]
                matched_count += 1
                print(f"  [✓] 匹配成功 ({best_match['score']}%) - {best_match['api'].get('流程名称', '')}")
                print(f"  依据: {best_match['reason']}\n")
                
                results.append({
                    'requirement': req,
                    'api': best_match['api'],
                    'score': best_match['score'],
                    'reason': best_match['reason'],
                    'inferred_flow': req.get('_inferred_flow', '')
                })
            else:
                print(f"  [✗] 未找到匹配\n")
                results.append({
                    'requirement': req,
                    'api': None,
                    'score': 0,
                    'reason': '',
                    'inferred_flow': req.get('_inferred_flow', '')
                })
        
        print("=" * 80)
        print(f"匹配完成: {matched_count}/{len(requirements)} 条需求找到匹配API\n")
        
        # 生成结果文件
        output_file = self.generate_output_xlsx(requirements, results, file_path)
        
        return output_file


def main():
    """主函数"""
    print("=" * 80)
    print("     API Document Retrieval V4 - 能率需求清单智能匹配")
    print("     优化: 开发分类自动推断流向 + 程序代号+流向匹配流程名称")
    print("=" * 80)
    print()
    
    if len(sys.argv) < 2:
        print("用法: python match_nenglv_v4.py <需求清单文件>")
        print("\n支持格式: .xls, .xlsx")
        print("\n示例:")
        print("  python match_nenglv_v4.py 需求评估整理/能率需求清单-bpm集成.xls")
        print("  python match_nenglv_v4.py 需求评估整理/能率需求清单-国家平台集成.xls")
        sys.exit(1)
    
    requirement_file = sys.argv[1]
    file_path = Path(requirement_file)
    
    if not file_path.exists():
        print(f"错误: 找不到文件 {requirement_file}")
        sys.exit(1)
    
    if file_path.suffix.lower() not in ['.xls', '.xlsx']:
        print("错误: 仅支持 Excel 文件 (.xls, .xlsx)")
        sys.exit(1)
    
    # 执行智能匹配
    matcher = NenglvMatcherV4()
    output_file = matcher.process_requirement_file(file_path)
    
    if output_file:
        print("=" * 80)
        print("智能匹配完成！")
        print("=" * 80)
        print(f"\n输出文件: {output_file}")
        print(f"在线文档: {ONLINE_DOC_URL}")
        print(f"\n请打开输出文件查看匹配结果！")


if __name__ == '__main__':
    main()
