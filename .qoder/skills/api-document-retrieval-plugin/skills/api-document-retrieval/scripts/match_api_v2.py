#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
API Document Retrieval V2 - 智能匹配增强版
新增规则：
1. 开发分类=710:客户端_推送接口 → 服务方T100，流向T100→XX
2. 客制类型=新增 → 匹配复用接口
3. 程序代号+流向 与接口清单的"流程名称"高度匹配
4. 匹配成功后追加到原需求清单
"""

import sys
import os
from pathlib import Path
from datetime import datetime
import re
import shutil

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("错误: 未安装 openpyxl，请运行: pip install openpyxl")
    sys.exit(1)

# 配置
WORKSPACE_DIR = Path(__file__).parent.parent.parent.parent.parent
API_MAPPING_FILE = WORKSPACE_DIR / "需求对照工具表" / "【2026】T100&OA类系统集成清单字段对照表.xlsx"
ONLINE_DOC_URL = "https://docs.qq.com/sheet/DSmdhek1TT3BPenRa?tab=BB08J2"

class IntelligentMatcher:
    """智能匹配器 - 支持需求清单Excel的智能匹配"""
    
    def __init__(self):
        self.api_data = {}  # 接口清单数据
        self.load_api_mapping()
    
    def load_api_mapping(self):
        """加载API对照清单 - 接口清单页签"""
        if not API_MAPPING_FILE.exists():
            print(f"警告: 找不到本地API对照清单文件: {API_MAPPING_FILE}")
            print(f"\n建议使用在线文档: {ONLINE_DOC_URL}")
            sys.exit(1)
        
        print(f"正在加载API对照清单: {API_MAPPING_FILE.name}")
        wb = openpyxl.load_workbook(API_MAPPING_FILE, data_only=True)
        
        # 查找"接口清单"页签
        interface_sheet = None
        for sheet_name in wb.sheetnames:
            if '接口清单' in sheet_name or 'interface' in sheet_name.lower():
                interface_sheet = wb[sheet_name]
                print(f"找到接口清单页签: {sheet_name}")
                break
        
        if not interface_sheet:
            print(f"未找到'接口清单'页签，使用第一个页签")
            interface_sheet = wb.active
        
        # 读取接口清单数据
        self.api_data['sheet_name'] = interface_sheet.title
        self.api_data['headers'] = []
        self.api_data['rows'] = []
        
        # 读取表头（第一行）
        for col_num, cell in enumerate(interface_sheet[1], 1):
            if cell.value:
                self.api_data['headers'].append({
                    'col': col_num,
                    'name': str(cell.value).strip()
                })
        
        # 读取所有行数据
        for row_num, row in enumerate(interface_sheet.iter_rows(min_row=2, values_only=False), 2):
            row_data = {}
            for header in self.api_data['headers']:
                cell = row[header['col'] - 1]
                row_data[header['name']] = cell.value if cell.value else ''
            
            row_data['_row_num'] = row_num
            row_data['_sheet'] = interface_sheet.title
            self.api_data['rows'].append(row_data)
        
        print(f"已加载 {len(self.api_data['rows'])} 条接口记录")
        print(f"字段: {[h['name'] for h in self.api_data['headers'][:10]]}\n")
    
    def extract_requirement_info(self, ws):
        """从需求清单中提取关键信息"""
        # 查找关键列
        headers = {}
        for col_num, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[str(cell.value).strip()] = col_num
        
        # 提取需求行
        requirements = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
            req_data = {'_row_num': row_num}
            
            # 提取开发分类
            dev_class_col = None
            for col_name, col_num in headers.items():
                if '开发分类' in col_name or '开发类别' in col_name:
                    dev_class_col = col_num
                    break
            
            if dev_class_col:
                req_data['开发分类'] = row[dev_class_col - 1].value if row[dev_class_col - 1].value else ''
            
            # 提取客制类型
            custom_type_col = None
            for col_name, col_num in headers.items():
                if '客制类型' in col_name or '定制类型' in col_name:
                    custom_type_col = col_num
                    break
            
            if custom_type_col:
                req_data['客制类型'] = row[custom_type_col - 1].value if row[custom_type_col - 1].value else ''
            
            # 提取程序代号
            prog_code_col = None
            for col_name, col_num in headers.items():
                if '程序代号' in col_name or '程序编号' in col_name:
                    prog_code_col = col_num
                    break
            
            if prog_code_col:
                req_data['程序代号'] = row[prog_code_col - 1].value if row[prog_code_col - 1].value else ''
            
            # 提取流向
            flow_dir_col = None
            for col_name, col_num in headers.items():
                if '流向' in col_name or '方向' in col_name:
                    flow_dir_col = col_num
                    break
            
            if flow_dir_col:
                req_data['流向'] = row[flow_dir_col - 1].value if row[flow_dir_col - 1].value else ''
            
            # 保存原始行数据
            req_data['_original_row'] = row
            req_data['_headers'] = headers
            
            requirements.append(req_data)
        
        return requirements, headers
    
    def match_interface(self, requirement):
        """根据需求匹配接口清单"""
        matches = []
        
        # 提取关键信息
        dev_class = str(requirement.get('开发分类', ''))
        custom_type = str(requirement.get('客制类型', ''))
        prog_code = str(requirement.get('程序代号', ''))
        flow_dir = str(requirement.get('流向', ''))
        
        # 规则1: 开发分类 = 710:客户端_推送接口
        if '710' in dev_class and '推送接口' in dev_class:
            # 服务方是T100，流向是T100→XX
            if not flow_dir or flow_dir.strip() == '':
                requirement['流向'] = 'T100→外部系统'
            
            # 规则2: 客制类型 = 新增，需要匹配复用接口
            if '新增' in custom_type:
                # 匹配逻辑：程序代号 + 流向 与 流程名称 高度匹配
                for api_row in self.api_data['rows']:
                    flow_name = str(api_row.get('流程名称', ''))
                    
                    # 计算匹配度
                    match_score = self.calculate_match_score(
                        prog_code, flow_dir, flow_name
                    )
                    
                    if match_score >= 70:  # 匹配度 >= 70%
                        matches.append({
                            'api': api_row,
                            'score': match_score,
                            'match_type': '流程名称匹配',
                            'reason': f"程序代号:{prog_code}, 流向:{flow_dir} → 流程名称:{flow_name}"
                        })
        
        # 按匹配度排序
        matches.sort(key=lambda x: x['score'], reverse=True)
        
        return matches
    
    def calculate_match_score(self, prog_code, flow_dir, flow_name):
        """计算匹配度分数"""
        score = 0
        
        # 1. 程序代号匹配（权重50%）
        if prog_code and prog_code in flow_name:
            score += 50
        
        # 2. 流向匹配（权重30%）
        if flow_dir and flow_dir in flow_name:
            score += 30
        elif 'T100' in flow_name and ('T100→' in flow_dir or 'T100' in flow_dir):
            score += 20
        
        # 3. 关键词匹配（权重20%）
        keywords = ['推送', '接口', '同步', '创建', '查询']
        for keyword in keywords:
            if keyword in prog_code and keyword in flow_name:
                score += 10
                break
        
        # 归一化到0-100
        return min(score, 100)
    
    def append_match_to_requirement(self, ws, req_row, matches, headers):
        """将匹配结果追加到需求清单"""
        row_num = req_row['_row_num']
        
        # 查找或创建匹配结果列
        match_cols = {}
        last_col = ws.max_column
        
        match_fields = [
            '匹配状态', '匹配接口', '流程名称', '接口方向', 
            '匹配度', '匹配依据', '参考位置', '在线文档'
        ]
        
        for field in match_fields:
            # 查找是否已存在该列
            found = False
            for col_num in range(1, last_col + 1):
                cell = ws.cell(row=1, column=col_num)
                if cell.value == field:
                    match_cols[field] = col_num
                    found = True
                    break
            
            # 不存在则创建新列
            if not found:
                last_col += 1
                match_cols[field] = last_col
                ws.cell(row=1, column=last_col, value=field)
        
        # 填充匹配结果
        if matches:
            best_match = matches[0]  # 取最佳匹配
            api = best_match['api']
            
            ws.cell(row=row_num, column=match_cols['匹配状态'], value='✅ 已匹配')
            ws.cell(row=row_num, column=match_cols['匹配接口'], value=api.get('接口名称', ''))
            ws.cell(row=row_num, column=match_cols['流程名称'], value=api.get('流程名称', ''))
            ws.cell(row=row_num, column=match_cols['接口方向'], value=api.get('接口方向', ''))
            ws.cell(row=row_num, column=match_cols['匹配度'], value=f"{best_match['score']}%")
            ws.cell(row=row_num, column=match_cols['匹配依据'], value=best_match['reason'])
            ws.cell(row=row_num, column=match_cols['参考位置'], 
                   value=f"{api.get('_sheet', '')} 行{api.get('_row_num', '')}")
            ws.cell(row=row_num, column=match_cols['在线文档'], value=ONLINE_DOC_URL)
        else:
            ws.cell(row=row_num, column=match_cols['匹配状态'], value='❌ 未匹配')
    
    def process_requirement_file(self, file_path):
        """处理需求清单文件"""
        print(f"\n📖 处理需求清单: {file_path.name}")
        
        # 加载需求清单
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        print(f"✓ 工作表: {ws.title}")
        print(f"✓ 行数: {ws.max_row - 1}")
        
        # 提取需求信息
        requirements, headers = self.extract_requirement_info(ws)
        print(f"✓ 提取到 {len(requirements)} 条需求\n")
        
        # 匹配接口
        match_count = 0
        for i, req in enumerate(requirements, 1):
            print(f"[{i}/{len(requirements)}] 匹配需求行 {req['_row_num']}...", end=' ')
            
            matches = self.match_interface(req)
            
            if matches:
                self.append_match_to_requirement(ws, req, matches, headers)
                match_count += 1
                print(f"✓ 匹配成功 (匹配度: {matches[0]['score']}%)")
            else:
                print("⚠ 未找到匹配")
        
        print(f"\n📊 匹配统计: {match_count}/{len(requirements)} 条需求匹配成功")
        
        # 保存结果
        output_file = file_path.parent / f"{file_path.stem}_已匹配{file_path.suffix}"
        
        # 备份原文件
        backup_file = file_path.parent / f"{file_path.stem}_备份_{datetime.now().strftime('%Y%m%d_%H%M%S')}{file_path.suffix}"
        shutil.copy2(file_path, backup_file)
        print(f"✓ 原文件已备份: {backup_file.name}")
        
        # 保存匹配结果
        wb.save(output_file)
        print(f"✓ 匹配结果已保存: {output_file.name}")
        print(f"✓ 文件大小: {output_file.stat().st_size / 1024:.2f} KB\n")
        
        return output_file


def main():
    """主函数"""
    print("=" * 80)
    print("              API Document Retrieval V2 - 智能匹配增强版")
    print("=" * 80)
    print()
    
    if len(sys.argv) < 2:
        print("用法: python match_api_v2.py <需求清单Excel文件>")
        print("\n示例:")
        print("  python match_api_v2.py 集成需求清单产出/xxx_标准化.xlsx")
        sys.exit(1)
    
    requirement_file = sys.argv[1]
    
    if not Path(requirement_file).exists():
        print(f"错误: 找不到文件 {requirement_file}")
        sys.exit(1)
    
    if not requirement_file.endswith(('.xlsx', '.xls')):
        print("错误: 仅支持 Excel 文件 (.xlsx, .xls)")
        sys.exit(1)
    
    # 执行智能匹配
    matcher = IntelligentMatcher()
    output_file = matcher.process_requirement_file(Path(requirement_file))
    
    print("=" * 80)
    print("智能匹配完成！")
    print("=" * 80)
    print(f"\n输出文件: {output_file}")
    print(f"打开文件查看匹配结果")
    print(f"在线文档: {ONLINE_DOC_URL}")


if __name__ == '__main__':
    main()
