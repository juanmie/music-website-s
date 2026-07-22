#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
API Document Retrieval - 自动匹配需求清单与API对照清单
根据需求清单中的关键字，从T100与OA类系统集成标准API对照清单中匹配对应的API接口和字段映射
"""

import sys
import os
from pathlib import Path
from datetime import datetime
import re

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("错误: 未安装 openpyxl，请运行: pip install openpyxl")
    sys.exit(1)

try:
    from docx import Document
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

# 配置
WORKSPACE_DIR = Path(__file__).parent.parent.parent.parent
API_MAPPING_FILE = WORKSPACE_DIR / "需求对照工具表" / "【2026】T100&OA类系统集成清单字段对照表.xlsx"
ONLINE_DOC_URL = "https://docs.qq.com/sheet/DSmdhek1TT3BPenRa?tab=BB08J2"

class KeywordExtractor:
    """关键字提取器"""
    
    def __init__(self):
        # 业务对象映射表
        self.business_objects = {
            '采购订单': ['PurchaseOrder', 'PO', '采购单'],
            '销售订单': ['SalesOrder', 'SO', '销售单'],
            '入库单': ['Receipt', 'GR', '入库'],
            '出库单': ['Delivery', 'GI', '出库'],
            '审批流程': ['Approval', 'Workflow', '审批'],
            '请假申请': ['LeaveRequest', '请假'],
            '报销申请': ['ExpenseRequest', '报销'],
            '物料': ['Material', 'Item', '物料'],
            '客户': ['Customer', '客户'],
            '供应商': ['Supplier', 'Vendor', '供应商'],
        }
        
        # 字段映射表
        self.field_mappings = {
            '单号': ['docno', 'doc_no', 'documentNo', 'orderNo'],
            '状态': ['status', 'state', 'approvalStatus'],
            '日期': ['date', 'createDate', 'docDate', 'approvalDate'],
            '申请人': ['requester', 'applicant', 'employee', 'emp_no'],
            '金额': ['amount', 'totalAmount', 'total_amt'],
            '数量': ['quantity', 'qty', '数量'],
            '备注': ['remark', 'comments', 'note', '备注'],
            '部门': ['department', 'dept', '部门'],
        }
    
    def extract_from_text(self, text):
        """从文本中提取关键字"""
        keywords = {
            'business_objects': [],
            'fields': [],
            'systems': [],
            'actions': [],
            'raw_keywords': []
        }
        
        # 提取业务对象
        for obj, synonyms in self.business_objects.items():
            if obj in text or any(syn in text for syn in synonyms):
                keywords['business_objects'].append(obj)
                keywords['raw_keywords'].append(obj)
        
        # 提取字段
        for field, synonyms in self.field_mappings.items():
            if field in text or any(syn.lower() in text.lower() for syn in synonyms):
                keywords['fields'].append(field)
                keywords['raw_keywords'].append(field)
        
        # 提取系统标识
        systems = ['T100', 'OA', 'ERP', 'WMS', 'SRM', 'MES', 'CRM']
        for system in systems:
            if system in text.upper():
                keywords['systems'].append(system)
                keywords['raw_keywords'].append(system)
        
        # 提取动作
        actions = ['创建', 'create', '更新', 'update', '删除', 'delete', 
                  '查询', 'query', '同步', 'sync', '审批', 'approve', '抛转']
        for action in actions:
            if action in text.lower():
                keywords['actions'].append(action)
                keywords['raw_keywords'].append(action)
        
        return keywords


class APIMatcher:
    """API匹配器"""
    
    def __init__(self):
        self.api_data = []
        self.load_api_mapping()
    
    def load_api_mapping(self):
        """加载API对照清单"""
        if not API_MAPPING_FILE.exists():
            print(f"⚠ 警告: 找不到本地API对照清单文件: {API_MAPPING_FILE}")
            print(f"\n💡 建议使用在线文档:")
            print(f"   {ONLINE_DOC_URL}")
            print(f"\n📥 下载指引:")
            print(f"   1. 打开上述链接")
            print(f"   2. 点击右上角「下载」按钮")
            print(f"   3. 选择 Excel 格式")
            print(f"   4. 保存到: {API_MAPPING_FILE.parent}")
            print(f"   5. 确保文件名: {API_MAPPING_FILE.name}")
            print(f"\n❓ 无法访问在线文档?")
            print(f"   - 检查网络连接")
            print(f"   - 确认腾讯文档账号登录")
            print(f"   - 联系文档管理员获取权限")
            sys.exit(1)
        
        print(f"正在加载API对照清单: {API_MAPPING_FILE.name}")
        wb = openpyxl.load_workbook(API_MAPPING_FILE, data_only=True)
        
        # 读取所有工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row_num, row in enumerate(ws.iter_rows(min_row=1, values_only=False), 1):
                row_data = {
                    'sheet': sheet_name,
                    'row': row_num,
                    'values': [cell.value for cell in row]
                }
                self.api_data.append(row_data)
        
        print(f"✓ 已加载 {len(self.api_data)} 行数据\n")
    
    def match_keywords(self, keywords):
        """根据关键字匹配API"""
        matches = {
            'high': [],
            'medium': [],
            'low': []
        }
        
        for api_row in self.api_data:
            row_text = ' '.join([str(v) for v in api_row['values'] if v])
            score = 0
            match_details = []
            
            # 高优先级：业务对象匹配
            for obj in keywords['business_objects']:
                if obj in row_text:
                    score += 50
                    match_details.append(f'业务对象: {obj}')
            
            # 中优先级：字段匹配
            for field in keywords['fields']:
                if field in row_text:
                    score += 20
                    match_details.append(f'字段: {field}')
            
            # 低优先级：系统标识匹配
            for system in keywords['systems']:
                if system in row_text:
                    score += 10
                    match_details.append(f'系统: {system}')
            
            # 动作匹配
            for action in keywords['actions']:
                if action in row_text.lower():
                    score += 15
                    match_details.append(f'动作: {action}')
            
            # 分类匹配结果
            if score >= 50:
                matches['high'].append({
                    'api': api_row,
                    'score': score,
                    'details': match_details
                })
            elif score >= 20:
                matches['medium'].append({
                    'api': api_row,
                    'score': score,
                    'details': match_details
                })
            elif score >= 10:
                matches['low'].append({
                    'api': api_row,
                    'score': score,
                    'details': match_details
                })
        
        # 按分数排序
        for level in matches:
            matches[level].sort(key=lambda x: x['score'], reverse=True)
        
        return matches


class ReportGenerator:
    """报告生成器"""
    
    def __init__(self):
        self.output_dir = WORKSPACE_DIR
    
    def generate_markdown_report(self, requirement_file, keywords, matches):
        """生成Markdown格式报告"""
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        date_str = datetime.now().strftime('%Y%m%d')
        
        report = []
        report.append("# API 匹配报告\n")
        report.append(f"生成时间: {timestamp}\n")
        
        # 需求信息
        report.append("## 📋 需求信息\n")
        report.append(f"- **需求文件**: {Path(requirement_file).name}")
        report.append(f"- **提取关键字**: {', '.join(keywords['raw_keywords'])}")
        report.append(f"- **业务对象**: {', '.join(keywords['business_objects'])}")
        report.append(f"- **关键字段**: {', '.join(keywords['fields'])}")
        report.append(f"- **涉及系统**: {', '.join(keywords['systems'])}")
        report.append(f"- **操作类型**: {', '.join(keywords['actions'])}\n")
        report.append(f"- **数据源**:")
        report.append(f"  - 本地文件: `需求对照工具表/【2026】T100&OA类系统集成清单字段对照表.xlsx`")
        report.append(f"  - 在线文档: {ONLINE_DOC_URL}\n")
        
        # 匹配统计
        report.append("## 📊 匹配统计\n")
        report.append(f"- 🔴 **高匹配度**: {len(matches['high'])} 个")
        report.append(f"- 🟡 **中匹配度**: {len(matches['medium'])} 个")
        report.append(f"- 🟢 **低匹配度**: {len(matches['low'])} 个\n")
        
        # 高匹配度结果
        if matches['high']:
            report.append("## 🔴 高匹配度 API\n")
            for i, match in enumerate(matches['high'][:10], 1):  # 最多显示10个
                api = match['api']
                report.append(f"### {i}. API 接口 (匹配度: {match['score']}%)\n")
                report.append(f"**参考位置**:")
                report.append(f"- 本地文件: 工作表 `{api['sheet']}` | 行号 `{api['row']}`")
                report.append(f"- 在线文档: {ONLINE_DOC_URL}\n")
                report.append(f"**匹配依据**: {', '.join(match['details'])}\n")
                
                # 显示字段映射（如果有的话）
                report.append("#### 字段映射")
                report.append("| 源字段 | 目标字段 | 类型 | 转换规则 |")
                report.append("|--------|---------|------|---------|")
                
                # 从行数据中提取字段信息（简化版）
                values = api['values']
                for j in range(0, min(len(values), 6), 2):
                    if j+1 < len(values) and values[j] and values[j+1]:
                        report.append(f"| {values[j]} | {values[j+1]} | - | - |")
                
                report.append("")
        
        # 中匹配度结果
        if matches['medium']:
            report.append("## 🟡 中匹配度 API\n")
            for i, match in enumerate(matches['medium'][:5], 1):
                api = match['api']
                report.append(f"### {i}. API 接口 (匹配度: {match['score']}%)\n")
                report.append(f"**参考位置**:")
                report.append(f"- 本地文件: 工作表 `{api['sheet']}` | 行号 `{api['row']}`")
                report.append(f"- 在线文档: {ONLINE_DOC_URL}\n")
                report.append(f"**匹配依据**: {', '.join(match['details'])}\n")
        
        # 低匹配度结果
        if matches['low']:
            report.append("## 🟢 低匹配度 API\n")
            report.append(f"共 {len(matches['low'])} 个低匹配度结果，建议补充需求描述以提高匹配度。\n")
        
        # 建议
        report.append("## 💡 建议\n")
        if not matches['high'] and not matches['medium']:
            report.append("- 需求描述可能不够详细，建议补充业务场景和字段信息")
            report.append("- 尝试使用更具体的接口名称或业务对象描述")
            report.append("- 明确接口方向（T100→OA 或 OA→T100）\n")
        else:
            report.append("- 建议人工复核高匹配度结果的准确性")
            report.append("- 重点关注字段映射的转换规则")
            report.append("- 参考对照清单原文档确认接口规范\n")
        
        report.append("---\n")
        report.append("*本报告由 API Document Retrieval 技能自动生成*")
        
        # 保存报告
        output_file = self.output_dir / f"API匹配报告_{date_str}.md"
        output_file.write_text('\n'.join(report), encoding='utf-8')
        
        return output_file


def extract_text_from_file(file_path):
    """从不同格式的文件中提取文本"""
    file_path = Path(file_path)
    text = ""
    
    if file_path.suffix == '.docx':
        if not HAS_DOCX:
            print("警告: 未安装 python-docx，无法读取 .docx 文件")
            return ""
        doc = Document(file_path)
        text = '\n'.join([p.text for p in doc.paragraphs if p.text.strip()])
    
    elif file_path.suffix in ['.xlsx', '.xls']:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        texts = []
        for ws in wb.sheetnames:
            for row in wb[ws].iter_rows(values_only=True):
                texts.extend([str(cell) for cell in row if cell])
        text = '\n'.join(texts)
    
    elif file_path.suffix in ['.txt', '.md']:
        text = file_path.read_text(encoding='utf-8')
    
    else:
        print(f"警告: 不支持的文件格式 {file_path.suffix}")
    
    return text


def main():
    """主函数"""
    print("=" * 80)
    print("                    API Document Retrieval")
    print("          T100与OA类系统集成API对照清单匹配工具")
    print("=" * 80)
    print()
    
    if len(sys.argv) < 2:
        print("用法: python match_api.py <需求清单文件路径>")
        print("\n支持的文件格式: .docx, .xlsx, .xls, .txt, .md")
        sys.exit(1)
    
    requirement_file = sys.argv[1]
    
    if not Path(requirement_file).exists():
        print(f"错误: 找不到文件 {requirement_file}")
        sys.exit(1)
    
    # 步骤 1: 提取文本
    print("📖 步骤 1: 提取需求清单内容...")
    text = extract_text_from_file(requirement_file)
    if not text:
        print("✗ 无法提取文本内容")
        sys.exit(1)
    print(f"✓ 已提取 {len(text)} 字符\n")
    
    # 步骤 2: 提取关键字
    print("🔍 步骤 2: 提取关键字...")
    extractor = KeywordExtractor()
    keywords = extractor.extract_from_text(text)
    
    print(f"  业务对象: {', '.join(keywords['business_objects'])}")
    print(f"  关键字段: {', '.join(keywords['fields'])}")
    print(f"  涉及系统: {', '.join(keywords['systems'])}")
    print(f"  操作类型: {', '.join(keywords['actions'])}")
    print(f"  总关键字: {len(keywords['raw_keywords'])}\n")
    
    if not keywords['raw_keywords']:
        print("⚠ 未提取到有效关键字，请检查需求清单内容")
        sys.exit(1)
    
    # 步骤 3: 匹配API
    print("🎯 步骤 3: 匹配API对照清单...")
    matcher = APIMatcher()
    matches = matcher.match_keywords(keywords)
    
    print(f"  🔴 高匹配度: {len(matches['high'])} 个")
    print(f"  🟡 中匹配度: {len(matches['medium'])} 个")
    print(f"  🟢 低匹配度: {len(matches['low'])} 个\n")
    
    # 步骤 4: 生成报告
    print("📝 步骤 4: 生成匹配报告...")
    generator = ReportGenerator()
    output_file = generator.generate_markdown_report(requirement_file, keywords, matches)
    
    print(f"✓ 报告已生成: {output_file}")
    print(f"✓ 文件大小: {output_file.stat().st_size / 1024:.2f} KB\n")
    
    print("=" * 80)
    print("✓ 匹配完成！")
    print("=" * 80)


if __name__ == '__main__':
    main()
